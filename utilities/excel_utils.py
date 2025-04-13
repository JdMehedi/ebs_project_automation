import openpyxl

def get_sheet(workbook, sheetname):
    if sheetname not in workbook.sheetnames:
        raise KeyError(f"Worksheet '{sheetname}' does not exist. Available sheets: {workbook.sheetnames}")
    return workbook[sheetname]

def get_row_count(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = get_sheet(workbook, sheetname)
    return sheet.max_row

def get_column_count(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = get_sheet(workbook, sheetname)
    return sheet.max_column

def read_data(file, sheetname, row_num, column_num):
    workbook = openpyxl.load_workbook(file)
    sheet = get_sheet(workbook, sheetname)
    return sheet.cell(row=row_num, column=column_num).value

def write_data(file, sheetname, row_num, column_num, data):
    workbook = openpyxl.load_workbook(file)
    sheet = get_sheet(workbook, sheetname)
    sheet.cell(row=row_num, column=column_num).value = data
    workbook.save(file)
